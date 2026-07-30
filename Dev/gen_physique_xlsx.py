# -*- coding: utf-8 -*-
"""生成体魄系统数值一览 Excel。数据来源于 PhysiqueDefine.cs / PhysiqueLgc.cs /
Hediff_PhysiqueDisplay.xml / ExerciseWork.cs / Trait_PhysiqueAptitudes.xml。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---------- 样式 ----------
HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SUB_FONT = Font(italic=True, size=9, color="808080")
STAGE_FILLS = {
    "虚弱": "F8CBAD", "一般": "FFE699", "健康": "C6E0B4",
    "强壮": "9DC3E6", "卓越": "B4A7D6",
}
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def apply_borders(ws, r0, r1, ncol, center=True):
    for r in range(r0, r1 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if center else LEFT


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, title, sub, ncol):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(row=2, column=1, value=sub).font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)


# ============================================================
# Sheet 1: 概览
# ============================================================
ws = wb.active
ws.title = "概览"
title_block(ws, "体魄系统 · 阶段总览", "体魄等级 0~20（可受特质扩展上限）；按等级映射为 5 个阶段", 6)
hdr = ["阶段", "等级区间", "Severity 区间", "核心定位", "肌肉劳损上限", "备注"]
for c, h in enumerate(hdr, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 6)
rows = [
    ["虚弱", "0 – 4", "0.00 – 0.24", "全面负面：移动/操作/工作/战斗均下降", 650, "易累、易拉伤"],
    ["一般", "5 – 7", "0.25 – 0.39", "基准状态，轻微负面", 1000, "普通人"],
    ["健康", "8 – 12", "0.40 – 0.64", "开始转正：命中/负重/呼吸提升", 1250, "肾上腺素惩罚解除起点(≥8)"],
    ["强壮", "13 – 16", "0.65 – 0.84", "全面强化：移动/操作/工作/战斗", 2000, "皮质醇增益、肾上腺素豁免起点(≥13)"],
    ["卓越", "17 – 20", "0.85 – 1.00", "顶级身体：各项大幅加成", 3000, "代价：饥饿速率翻倍"],
]
for i, row in enumerate(rows):
    r = 5 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=STAGE_FILLS[row[0]])
apply_borders(ws, 5, 9, 6)
set_widths(ws, [8, 12, 14, 34, 12, 30])

# ============================================================
# Sheet 2: 阶段 → 身体属性（XML 实际生效）
# ============================================================
ws = wb.create_sheet("阶段-身体属性(生效)")
title_block(ws, "阶段 → 身体属性修正（游戏内真实生效）",
            "来源 Hediff_PhysiqueDisplay.xml。offset=加减法，factor=乘法。空=无修正", 6)
hdr = ["属性", "类型", "虚弱", "一般", "健康", "强壮", "卓越"]
# 重建为 7 列
for c, h in enumerate(hdr, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 7)
data = [
    ["移动 Moving", "offset", "-0.10", "-0.03", "", "+0.03", "+0.08"],
    ["操作 Manipulation", "offset", "-0.08", "-0.03", "", "+0.03", "+0.08"],
    ["呼吸 Breathing", "offset", "-0.05", "", "+0.03", "+0.05", "+0.15"],
    ["近战命中 MeleeHitChance", "offset", "-0.10", "", "+0.10", "+0.10", "+0.20"],
    ["近战闪避 MeleeDodgeChance", "offset", "-0.20", "-0.10", "+0.10", "+0.20", "+0.30"],
    ["射击精度 ShootingAccuracyPawn", "offset", "-0.10", "", "+0.10", "+0.10", "+0.20"],
    ["近战伤害 MeleeDamageFactor", "factor", "0.90", "", "1.05", "1.10", "1.15"],
    ["移动速度 MoveSpeed", "factor", "0.90", "", "", "1.03", "1.08"],
    ["全局工作速度 WorkSpeedGlobal", "factor", "0.90", "0.95", "", "1.03", "1.08"],
    ["负重 CarryingCapacity", "factor", "0.90", "0.97", "1.25", "1.65", "2.25"],
    ["体重 Mass", "factor", "0.80", "", "1.10", "1.20", "1.25"],
    ["饥饿速率 hungerRateFactorOffset", "offset", "-0.20", "", "+0.20", "+0.60", "+1.00"],
]
for i, row in enumerate(data):
    r = 5 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
apply_borders(ws, 5, 4 + len(data), 7)
# 阶段列上色
stage_cols = {3: "虚弱", 4: "一般", 5: "健康", 6: "强壮", 7: "卓越"}
for c, name in stage_cols.items():
    ws.cell(row=4, column=c).fill = PatternFill("solid", fgColor=STAGE_FILLS[name])
    ws.cell(row=4, column=c).font = Font(bold=True, color="000000")
set_widths(ws, [30, 8, 9, 9, 9, 9, 9])

# ============================================================
# Sheet 3: 阶段 → 激素/劳损系统内部修正（C# 逻辑）
# ============================================================
ws = wb.create_sheet("阶段-系统内部修正")
title_block(ws, "阶段 → 激素 / 肌肉劳损系统内部系数",
            "来源 PhysiqueLgc.cs / PhysiqueDefine.cs，供 mod 内部逻辑使用（与上表用途不同）", 6)
hdr = ["项目", "虚弱", "一般", "健康", "强壮", "卓越"]
for c, h in enumerate(hdr, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 6)
data = [
    ["战斗命中加成 GetPhysiqueBonus", "0.90", "1.00", "1.10", "1.10", "1.20"],
    ["工作效率(内部) GetWorkEfficiency", "0.90", "0.95", "1.00", "1.03", "1.08"],
    ["代谢率 GetMetabolicRate", "0.95", "1.00", "1.03", "1.05", "1.15"],
    ["饥饿/食欲倍率 GetHungerRate", "0.80", "1.00", "1.20", "1.60", "2.00"],
    ["肌肉劳损上限 MuscleStrainMax", "650", "1000", "1250", "2000", "3000"],
    ["劳损扣减倍率 ConsumeMult", "2.00", "1.00", "0.75", "0.50", "0.30"],
    ["拉伤概率倍率 StrainChanceMult", "1.25", "1.00", "0.75", "0.50", "0.25"],
    ["劳损恢复倍率 StrainRecoveryMult", "0.90", "1.00", "1.25", "2.00", "3.00"],
]
for i, row in enumerate(data):
    r = 5 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
apply_borders(ws, 5, 4 + len(data), 6)
for c, name in {2: "虚弱", 3: "一般", 4: "健康", 5: "强壮", 6: "卓越"}.items():
    ws.cell(row=4, column=c).fill = PatternFill("solid", fgColor=STAGE_FILLS[name])
    ws.cell(row=4, column=c).font = Font(bold=True, color="000000")
set_widths(ws, [34, 10, 10, 10, 10, 10])

# ============================================================
# Sheet 4: 与等级线性相关的激素修正
# ============================================================
ws = wb.create_sheet("等级线性修正")
title_block(ws, "随等级线性变化的激素修正",
            "公式基于 PhysiqueMaxLevel=20；恢复加成 1.0→1.5，伤害减免 1.0→0.5", 5)
hdr = ["项目", "公式", "等级0", "等级10", "等级20"]
for c, h in enumerate(hdr, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 5)


def recov(lv):
    return round(1 + (lv - 1) / 19 * 0.5, 3)


def dmgred(lv):
    return round(1 - (lv - 1) / 19 * 0.5, 3)


data = [
    ["激素恢复加成 RecoveryBonus", "1 + (lv-1)/19 × 0.5", recov(0), recov(10), recov(20)],
    ["激素伤害减免 DamageReduction", "1 - (lv-1)/19 × 0.5", dmgred(0), dmgred(10), dmgred(20)],
]
for i, row in enumerate(data):
    r = 5 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
apply_borders(ws, 5, 4 + len(data), 5)
# 阈值型修正
r0 = 4 + len(data) + 2
ws.cell(row=r0, column=1, value="阈值型修正（分段常量）").font = Font(bold=True, color="1F3864")
r0 += 1
hdr2 = ["项目", "条件", "取值", "说明", ""]
for c, h in enumerate(hdr2, 1):
    ws.cell(row=r0, column=c, value=h)
style_header(ws, r0, 4)
thr = [
    ["肾上腺素修正 AdrenalineModifier", "体魄 < 8", "0.5", "负面影响加重"],
    ["", "体魄 ≥ 8", "1.0", "正常"],
    ["肾上腺素豁免 IsAdrenalineExempt", "体魄 ≥ 13", "豁免", "不受惩罚"],
    ["皮质醇修正 CortisolModifier", "体魄 < 8", "0.5", "上升快/下降慢"],
    ["", "8 ≤ 体魄 < 13", "1.0", "正常"],
    ["", "体魄 ≥ 13", "1.2", "积聚慢/消退快"],
]
for i, row in enumerate(thr):
    r = r0 + 1 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
apply_borders(ws, r0 + 1, r0 + len(thr), 4, center=False)
set_widths(ws, [32, 20, 12, 24, 4])

# ============================================================
# Sheet 5: 成长 —— 劳作/锻炼获取经验
# ============================================================
ws = wb.create_sheet("成长-经验获取")
title_block(ws, "体魄经验获取（劳作 & 锻炼）",
            "劳作干活按次给经验并可能拉伤；锻炼点专门练体魄。工作分档：重活/中活/轻活", 5)
hdr = ["行为", "经验XP", "劳损扣减", "拉伤概率", "分档/说明"]
for c, h in enumerate(hdr, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 5)
data = [
    ["挖矿 Mining", 100, 15, "6%", "重活"],
    ["深钻 DeepDrill", 60, 15, "4%", "重活"],
    ["砍树 TreeCut", 50, 15, "3%", "重活"],
    ["挖树桩 ExtractTree", 50, 15, "3%", "重活"],
    ["完成建造 FinishFrame", 40, 15, "3%", "重活"],
    ["打磨墙/地板 Smooth", 40, 15, "3%", "重活"],
    ["拆除 Deconstruct", 30, 15, "2%", "重活"],
    ["拆卸 Uninstall", 30, 15, "2%", "重活"],
    ["搬运 Haul", 25, 13, "1%", "中活"],
    ["宰杀 Butcher", 25, 13, "3%", "中活"],
    ["打猎 Hunt", 30, 13, "2%", "中活"],
    ["修理 Repair", 25, 13, "1%", "中活"],
    ["收割 Harvest", 25, 6, "1%", "轻活"],
    ["割草 PlantCut", 8, 6, "1%", "轻活"],
    ["播种 Sow", 10, 6, "1%", "轻活"],
    ["移植 Replant", 10, 6, "1%", "轻活"],
]
for i, row in enumerate(data):
    r = 5 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
apply_borders(ws, 5, 4 + len(data), 5)
set_widths(ws, [22, 10, 10, 10, 20])

# 锻炼参数块
r0 = 5 + len(data) + 1
ws.cell(row=r0, column=1, value="锻炼点（Hormones_ExerciseSpot）参数").font = Font(bold=True, color="1F3864")
r0 += 1
ex = [
    ["单次锻炼时长", "5000 tick", "约 2 游戏小时"],
    ["每 tick 体魄经验", "0.075 XP", "满疗程约 375 XP（会因中途停止而减少）"],
    ["每 tick 劳损扣减", "0.08 × 消耗倍率", "一般体魄满疗程约扣 400（≈40% 储备）"],
    ["最低体力门槛", "35%", "劳损储备低于上限35%无法开始/中途停止"],
]
hdr3 = ["项目", "数值", "说明"]
for c, h in enumerate(hdr3, 1):
    ws.cell(row=r0, column=c, value=h)
style_header(ws, r0, 3)
for i, row in enumerate(ex):
    r = r0 + 1 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
apply_borders(ws, r0 + 1, r0 + len(ex), 3, center=False)

# ============================================================
# Sheet 6: 特质 & 背景偏移
# ============================================================
ws = wb.create_sheet("特质偏移")
title_block(ws, "特质对体魄等级的偏移（skillGains.Physique）",
            "来源 Trait_PhysiqueAptitudes.xml；直接加到体魄技能等级上", 3)
hdr = ["特质", "档位", "体魄偏移"]
for c, h in enumerate(hdr, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 3)
data = [
    ["Bloodlust 嗜血", "-", "+3"],
    ["Brawler 斗士", "-", "+3"],
    ["Immunity 强免疫", "degree 1", "+3"],
    ["SpeedOffset 敏捷", "degree 1", "+3"],
    ["Tough 坚韧", "-", "+2"],
    ["Nimble 灵巧", "-", "+2"],
    ["QuickSleeper 快速睡眠", "-", "+1"],
    ["Wimp 懦弱", "-", "-1"],
    ["Delicate 娇弱", "-", "-2"],
    ["NaturalMood 天生抑郁", "degree -2", "-2"],
    ["Immunity 弱免疫", "degree -1", "-3"],
]
for i, row in enumerate(data):
    r = 5 + i
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        if c == 3:
            cell.font = Font(color="C00000" if str(v).startswith("-") else "006100", bold=True)
apply_borders(ws, 5, 4 + len(data), 3)
set_widths(ws, [26, 14, 10])
ws.cell(row=5 + len(data) + 1, column=1,
        value="注：背景故事(Backstory_Physique.xml)也会通过 skillGains 追加偏移；最终等级 = 技能等级 + 特质/背景偏移，并 Clamp 到 [0, 20+特质上限偏移]。").font = SUB_FONT

# ============================================================
# Sheet 7: 体魄日常衰减（用进废退，2026-07-30）
# ============================================================
ws = wb.create_sheet("日常衰减")
title_block(ws, "体魄日常衰减（用进废退）",
            "每游戏日(60000tick)结算一次；当天有过体力劳作或锻炼则免于衰减，闲置才按阶段扣 Physique 经验", 4)
hdr = ["阶段", "等级", "日衰减 XP/天", "说明"]
for c, h in enumerate(hdr, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 4)
decay = [
    ["虚弱", "0 – 4", 0, "保底，不惩罚新手/伤员"],
    ["一般", "5 – 7", 2, "极缓退步"],
    ["健康", "8 – 12", 4, "缓慢退步"],
    ["强壮", "13 – 16", 8, "需偶尔维护"],
    ["卓越", "17 – 20", 14, "维护成本最高，约2-3天掉1级"],
]
for i, row in enumerate(decay):
    r = 5 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    for c in range(1, 5):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=STAGE_FILLS[row[0]])
apply_borders(ws, 5, 4 + len(decay), 4)
set_widths(ws, [10, 12, 14, 34])

r0 = 5 + len(decay) + 1
rules = [
    ["结算周期", "60000 tick", "1 游戏日累计满触发一次"],
    ["活动豁免", "劳作 或 锻炼", "当天发生任一即免衰减；标记于下个周期重置"],
    ["作用对象", "仅类人", "复用 IsHormoneSubject，动物/机械体跳过"],
    ["虚弱保底", "衰减=0", "等级<5 不衰减，防止练废后继续掉"],
    ["玩家可调总倍率", "0 ~ 3（默认1）", "PhysiqueDecayGlobalMult，调0=完全关闭"],
    ["实现方式", "Learn(-xp, direct)", "负经验直扣，内部处理掉级；不触发学习速率修正"],
]
ws.cell(row=r0, column=1, value="机制要点").font = Font(bold=True, color="1F3864")
r0 += 1
hdr2 = ["项目", "取值", "说明"]
for c, h in enumerate(hdr2, 1):
    ws.cell(row=r0, column=c, value=h)
style_header(ws, r0, 3)
for i, row in enumerate(rules):
    r = r0 + 1 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
apply_borders(ws, r0 + 1, r0 + len(rules), 3, center=False)

# ============================================================
# Sheet 8: 体魄热情倍率覆盖（2026-07-30）
#   只对 Physique 技能覆盖原版 passion 学习倍率。
# ============================================================
ws = wb.create_sheet("热情倍率覆盖")
title_block(ws, "体魄技能 · 热情学习倍率覆盖",
            "仅对 Physique 技能生效；Prefix 补偿 SkillRecord.Learn 的 xp，先除原版倍率再乘自定义倍率。负经验(衰减)不受影响。", 4)
hdr = ["热情", "原版倍率", "覆盖后倍率", "说明"]
for c, h in enumerate(hdr, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 4)
passion = [
    ["无 (None)", "×0.333", "×1.00", "无热情也按满速学习，不再打3折"],
    ["好奇 (Minor)", "×1.00", "×1.10", "略高于原版"],
    ["狂热 (Major)", "×1.50", "×1.20", "压缩狂热优势，避免体魄暴涨"],
]
for i, row in enumerate(passion):
    r = 5 + i
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
apply_borders(ws, 5, 4 + len(passion), 4)
set_widths(ws, [14, 12, 14, 40])
ws.cell(row=5 + len(passion) + 1, column=1,
        value="注：原版倍率取自 RimWorld 1.6 SkillDef.json 的 PassionsData.LearningFactor（无 0.333 / 好奇 1.0 / 狂热 1.5）。"
              "实现于 HormonesLogic.cs 的 SkillRecord_Learn_Physique_Patch。").font = SUB_FONT

# 冻结首行标题
for sheet in wb.worksheets:
    sheet.freeze_panes = "A5"

out = r"D:\RimMods\Rim-Hormones\RimHormones\Dev\体魄系统数值一览.xlsx"
wb.save(out)
print("SAVED", out)
